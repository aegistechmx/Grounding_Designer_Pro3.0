"""
Professional Excel Export for Grounding Designer Pro
Uses openpyxl to generate engineering-grade Excel reports
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Load data from JSON
with open("data.json", "r") as f:
    data = json.load(f)

results = data.get("results", {})

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Grounding Results"

# Define styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

data_font = Font(size=11)
data_alignment = Alignment(horizontal="left", vertical="center")

border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add header row
headers = ["Parameter", "Value", "Unit"]
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# Add data rows
data_rows = [
    ["Grid Resistance (Rg)", results.get("Rg", 0), "Ω"],
    ["GPR", results.get("GPR", 0), "V"],
    ["Touch Voltage (Em)", results.get("Em", 0), "V"],
    ["Step Voltage (Es)", results.get("Es", 0), "V"],
    ["Permissible Touch", results.get("Etouch70", 0), "V"],
    ["Permissible Step", results.get("Estep70", 0), "V"],
    ["Grid Current (Ig)", results.get("Ig", 0), "A"],
    ["Fault Current", results.get("faultCurrent", 0), "A"],
    ["Fault Duration", results.get("faultDuration", 0), "s"],
    ["Soil Resistivity", results.get("soilResistivity", 0), "Ω·m"],
    ["Surface Layer Resistivity", results.get("surfaceLayer", 0), "Ω·m"],
    ["Surface Depth", results.get("surfaceDepth", 0), "m"],
    ["Grid Length", results.get("gridLength", 0), "m"],
    ["Grid Width", results.get("gridWidth", 0), "m"],
    ["Number of Conductors X", results.get("numParallel", 0), ""],
    ["Number of Conductors Y", results.get("numParallelY", 0), ""],
    ["Number of Rods", results.get("numRods", 0), ""],
    ["Rod Length", results.get("rodLength", 0), "m"],
    ["Grid Depth", results.get("gridDepth", 0), "m"],
]

for row_idx, row_data in enumerate(data_rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = border

# Add compliance status
ws.append([])
ws.append(["Compliance Status"])
ws.append(["Complies with IEEE 80", "YES" if results.get("complies") else "NO"])
ws.append(["Touch Voltage Safe", "YES" if results.get("touchSafe70") else "NO"])
ws.append(["Step Voltage Safe", "YES" if results.get("stepSafe70") else "NO"])

# Style compliance section
for row in range(len(data_rows) + 3, len(data_rows) + 7):
    for col in range(1, 3):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True if row == len(data_rows) + 3 else False, size=11)
        cell.border = border

# Add metadata
ws.append([])
ws.append(["Report Information"])
ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
ws.append(["Standard", "IEEE 80-2013"])

# Adjust column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 10

# Save workbook
wb.save("report.xlsx")
print("Excel generated successfully: report.xlsx")
